#!/usr/bin/env python3
"""
한국어(KR) 검수 확정본을 담는 'Teacher' 테이블(kr_teacher) 구축.

두 엑셀(신규 검수본 vs 기준본)의 Words_main KR 컬럼을 비교해,
'바뀐 text_id 만' kr_teacher 에 INSERT 한다. (기본은 빈 테이블 — 검수로 확정된 것만 쌓임)
game_texts 원본은 건드리지 않는다.

사용:
  PGPASSWORD='somr#' python scripts/load_kr_teacher.py \
      "여신키우기_텍스트_0722(초안 검수 1022).xlsx" 여신키우기_텍스트_0720.xlsx
"""
import os
import sys
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

KR_COL = 6  # Words_main 의 韩语 컬럼 인덱스(0-based)


def load_kr(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Words_main"]
    d = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        idv = row[0]
        if idv is None or str(idv).strip() == "":
            continue
        try:
            tid = int(float(str(idv).strip()))
        except ValueError:
            continue
        v = row[KR_COL] if KR_COL < len(row) else None
        d[tid] = "" if v is None else str(v).strip()
    return d


def main():
    if len(sys.argv) < 3:
        print("usage: load_kr_teacher.py <new.xlsx> <base.xlsx>")
        sys.exit(2)
    new_path, base_path = sys.argv[1], sys.argv[2]
    source = os.path.basename(new_path)

    new = load_kr(new_path)
    base = load_kr(base_path)
    changed = [
        (tid, new[tid], base.get(tid, ""))
        for tid in sorted(new)
        if new[tid] != base.get(tid, "")
    ]
    print(f"[kr_teacher] KR 변경 {len(changed)}건 (source={source})")

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "192.168.0.2"),
        dbname=os.environ.get("PGDATABASE", "ThePlayPlus"),
        user=os.environ.get("PGUSER", "gino"),
        password=os.environ["PGPASSWORD"],
    )
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS kr_teacher (
                    text_id  BIGINT PRIMARY KEY,
                    kr       TEXT NOT NULL,   -- 검수 확정 한국어 (teacher)
                    base_kr  TEXT,            -- 수정 전 값 (참고)
                    source   TEXT,            -- 출처(검수 파일)
                    added_at TIMESTAMPTZ NOT NULL DEFAULT now()
                );
                COMMENT ON TABLE kr_teacher IS '한국어 검수 확정본(Teacher) — 검수로 바뀐 것만 적재';
            """)
            rows = [(t, kr, b, source) for (t, kr, b) in changed]
            execute_values(
                cur,
                """INSERT INTO kr_teacher (text_id, kr, base_kr, source) VALUES %s
                   ON CONFLICT (text_id) DO UPDATE SET
                     kr = EXCLUDED.kr, base_kr = EXCLUDED.base_kr,
                     source = EXCLUDED.source, added_at = now()""",
                rows,
            )
            cur.execute("SELECT count(*) FROM kr_teacher")
            n = cur.fetchone()[0]
        conn.commit()
        print(f"[kr_teacher] ✅ 적재 완료. 총 {n}행")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
