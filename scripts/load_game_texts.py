#!/usr/bin/env python3
"""
여신키우기_텍스트_*.xlsx 의 Words_main 시트를 PostgreSQL(ThePlayPlus) 테이블로 적재.

테이블: game_texts (text_id PK, 언어별 컬럼 + note + char_limit)
멱등: 실행할 때마다 CREATE IF NOT EXISTS 후 TRUNCATE + 재적재.

사용:
  pip install openpyxl psycopg2-binary
  PGPASSWORD='somr#' python scripts/load_game_texts.py 여신키우기_텍스트_0720.xlsx
"""
import os
import sys
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

XLSX = sys.argv[1] if len(sys.argv) > 1 else "여신키우기_텍스트_0720.xlsx"
SHEET = "Words_main"
HEADER_ROWS = 4  # 상단 4행은 헤더/타입 메타 → 스킵

# (엑셀 컬럼 index, DB 컬럼명)
COLS = [
    (0, "text_id"),
    (1, "note"),
    (2, "cn"), (3, "en"), (4, "jp"), (5, "cnt"),
    (6, "kr"), (7, "vn"), (8, "pt"), (9, "th"), (10, "my"),
    (11, "char_limit"),
]
DB_COLS = [c for _, c in COLS]

DDL = """
CREATE TABLE IF NOT EXISTS game_texts (
    text_id    BIGINT PRIMARY KEY,   -- 文本ID
    note       TEXT,                 -- 文本备注说明
    cn         TEXT,                 -- 中文 (원문)
    en         TEXT,                 -- 英文
    jp         TEXT,                 -- 日语
    cnt        TEXT,                 -- 繁体中文
    kr         TEXT,                 -- 韩语
    vn         TEXT,                 -- 越南语
    pt         TEXT,                 -- 葡萄牙语
    th         TEXT,                 -- 泰语
    my         TEXT,                 -- 马来西亚语
    char_limit TEXT                  -- 字符限制
);
COMMENT ON TABLE game_texts IS '여신키우기 다국어 텍스트 (Words_main)';
"""


def cell_to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def main():
    print(f"[load] 엑셀 읽는 중: {XLSX} / 시트 {SHEET}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[SHEET]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < HEADER_ROWS:
            continue
        idv = row[0]
        if idv is None or str(idv).strip() == "":
            continue
        try:
            text_id = int(float(str(idv).strip()))
        except ValueError:
            print(f"  경고: {i}행 ID 정수 아님 → 스킵: {idv!r}")
            continue
        rec = [text_id]
        for idx, _name in COLS[1:]:
            rec.append(cell_to_text(row[idx]) if idx < len(row) else None)
        rows.append(rec)
    print(f"[load] 적재 대상 행: {len(rows)}")

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "192.168.0.2"),
        port=int(os.environ.get("PGPORT", "5432")),
        dbname=os.environ.get("PGDATABASE", "ThePlayPlus"),
        user=os.environ.get("PGUSER", "gino"),
        password=os.environ["PGPASSWORD"],
    )
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            cur.execute(DDL)
            cur.execute("TRUNCATE game_texts;")
            execute_values(
                cur,
                f"INSERT INTO game_texts ({', '.join(DB_COLS)}) VALUES %s",
                rows,
                page_size=1000,
            )
            cur.execute("SELECT count(*) FROM game_texts;")
            n = cur.fetchone()[0]
        conn.commit()
        print(f"[load] ✅ 완료. game_texts 행 수: {n}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
