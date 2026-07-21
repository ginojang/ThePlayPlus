#!/usr/bin/env python3
"""
여신키우기_텍스트_*.xlsx 의 나머지 21개 시트를 PostgreSQL(ThePlayPlus)에 1:1 적재.

Words_main 은 이미 game_texts 로 별도 적재됨(scripts/load_game_texts.py) → 여기선 제외.

전략(충실 덤프):
  - 시트마다 테이블 gt_<slug> 생성. 모든 셀을 TEXT 로, 헤더행 포함 전 행을 그대로 적재.
  - row_idx(0-based) 로 원본 순서 보존. 컬럼은 실제 사용된 최대 열까지 c01..cNN.
  - gt_sheets 레지스트리에 (테이블명 ↔ 원본 시트명, 행/열 수) 기록.
  - 멱등: 대상 테이블 DROP 후 재생성.

사용:
  PGPASSWORD='somr#' python scripts/load_all_sheets.py 여신키우기_텍스트_0720.xlsx
"""
import os
import sys
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values

XLSX = sys.argv[1] if len(sys.argv) > 1 else "여신키우기_텍스트_0720.xlsx"
EXCLUDE = {"Words_main"}  # 이미 game_texts 로 적재됨

# 원본 시트명 → ASCII 테이블명(gt_ 접두)
SLUG = {
    "修改记录": "gt_changelog",
    "目录": "gt_toc",
    "主角技能描述": "gt_hero_skill_desc",
    "Sheet22": "gt_sheet22",
    "id分配": "gt_id_alloc",
    "Sheet2": "gt_sheet2",
    "技能名称": "gt_skill_names",
    "Sheet3": "gt_sheet3",
    "Sheet4": "gt_sheet4",
    "技能配置辅助": "gt_skill_config_aux",
    "Sheet5": "gt_sheet5",
    "Sheet6": "gt_sheet6",
    "韩越": "gt_kr_vn",
    "韩越待翻译": "gt_kr_vn_todo",
    "1106英语越语补充": "gt_1106_en_vn_supp",
    "1108内部翻译": "gt_1108_internal",
    "1108外部韩越": "gt_1108_ext_kr_vn",
    "1110": "gt_1110",
    "1212日语翻译": "gt_1212_jp",
    "Sheet17": "gt_sheet17",
    "250606繁中翻译": "gt_250606_cnt",
}


def cell_to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def used_cols(ws):
    """실제 값이 존재하는 최대 열 수 (뒤쪽 완전 빈 열 제거)."""
    n = 0
    for row in ws.iter_rows(values_only=True):
        for j in range(len(row) - 1, -1, -1):
            if cell_to_text(row[j]) is not None:
                n = max(n, j + 1)
                break
    return n


def main():
    print(f"[load-all] 엑셀: {XLSX}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "192.168.0.2"),
        port=int(os.environ.get("PGPORT", "5432")),
        dbname=os.environ.get("PGDATABASE", "ThePlayPlus"),
        user=os.environ.get("PGUSER", "gino"),
        password=os.environ["PGPASSWORD"],
    )
    conn.autocommit = False
    summary = []
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS gt_sheets (
                    table_name TEXT PRIMARY KEY,
                    sheet_name TEXT NOT NULL,
                    n_rows INT,
                    n_cols INT
                );
                COMMENT ON TABLE gt_sheets IS '여신키우기 엑셀 시트 ↔ 테이블 매핑';
            """)

            for name in wb.sheetnames:
                if name in EXCLUDE:
                    continue
                table = SLUG.get(name)
                if not table:
                    print(f"  경고: slug 미정의 시트 스킵: {name!r}")
                    continue
                ws = wb[name]
                ncols = used_cols(ws)
                if ncols == 0:
                    print(f"  - {name} → {table}: 빈 시트, 스킵")
                    continue

                cols = [f"c{j+1:02d}" for j in range(ncols)]
                coldefs = ",\n  ".join(f"{c} TEXT" for c in cols)
                cur.execute(f'DROP TABLE IF EXISTS {table};')
                cur.execute(
                    f'CREATE TABLE {table} (\n  row_idx INT PRIMARY KEY,\n  {coldefs}\n);'
                )
                cur.execute(
                    "COMMENT ON TABLE %s IS %%s" % table, (f"여신키우기 시트 '{name}' 원본 덤프",)
                )

                data = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    vals = [cell_to_text(row[j]) if j < len(row) else None for j in range(ncols)]
                    if all(v is None for v in vals):
                        continue  # 완전 빈 행 스킵(순서 위해 row_idx 는 원본 유지)
                    data.append([i] + vals)

                if data:
                    execute_values(
                        cur,
                        f"INSERT INTO {table} (row_idx, {', '.join(cols)}) VALUES %s",
                        data,
                        page_size=1000,
                    )
                cur.execute(
                    "INSERT INTO gt_sheets (table_name, sheet_name, n_rows, n_cols) "
                    "VALUES (%s,%s,%s,%s) ON CONFLICT (table_name) DO UPDATE SET "
                    "sheet_name=EXCLUDED.sheet_name, n_rows=EXCLUDED.n_rows, n_cols=EXCLUDED.n_cols",
                    (table, name, len(data), ncols),
                )
                summary.append((name, table, len(data), ncols))
                print(f"  - {name:<18} → {table:<22} rows={len(data):>6} cols={ncols}")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"\n[load-all] ✅ {len(summary)}개 시트 적재 완료")


if __name__ == "__main__":
    main()
